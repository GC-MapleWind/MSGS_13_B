import argparse
import asyncio
import os
from pathlib import Path

import openpyxl
from sqlalchemy import select

from src.database import async_session, init_db
from src.models.team import TeamMember, TeamMessage


ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_DATA_DIR = ROOT_DIR / "13기 메생결산"
DATA_DIR = Path(os.environ.get("INIT_DATA_DIR", str(DEFAULT_DATA_DIR)))

XLSX_PATH = DATA_DIR / "메생결산 운영팀 한마디(응답).xlsx"
IMAGE_DIR = DATA_DIR / "운영팀한마디"
IMAGE_URL_PREFIX = "/static/settlements/운영팀한마디"


def _find_image_url(name: str) -> str | None:
    for ext in (".png", ".jpg", ".jpeg", ".PNG", ".JPG", ".JPEG", ".webp", ".gif"):
        candidate = IMAGE_DIR / f"{name}{ext}"
        if candidate.exists():
            return f"{IMAGE_URL_PREFIX}/{candidate.name}"
    return None


def _load_rows() -> list[dict[str, str | None]]:
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"XLSX not found: {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[wb.sheetnames[0]]

    by_name: dict[str, dict[str, str | None]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, name, title, content, _photo = row[:5]
        if not name:
            continue

        name_text = str(name).strip()
        if not name_text:
            continue

        by_name[name_text] = {
            "name": name_text,
            "role": "운영팀",
            "title": str(title).strip() if title else "운영팀 한마디",
            "content": str(content).strip() if content else "",
            "image_url": _find_image_url(name_text),
        }

    return list(by_name.values())


async def seed_team_messages(prune_missing: bool = False) -> None:
    await init_db()
    rows = _load_rows()
    names = {r["name"] for r in rows}

    async with async_session() as db:
        existing_members = (await db.execute(select(TeamMember))).scalars().all()
        members_by_name = {m.name: m for m in existing_members}

        removed = 0
        if prune_missing:
            for member in existing_members:
                if member.name in names:
                    continue
                msg = (
                    await db.execute(
                        select(TeamMessage).where(TeamMessage.member_id == member.id)
                    )
                ).scalar_one_or_none()
                if msg is not None:
                    await db.delete(msg)
                await db.delete(member)
                removed += 1

        upserted = 0
        for row in rows:
            name = str(row["name"])
            role = str(row["role"])
            title = str(row["title"])
            content = str(row["content"])
            image_url = row["image_url"]

            member = members_by_name.get(name)
            if member is None:
                member = TeamMember(
                    name=name,
                    role=role,
                    profile_img_url=image_url,
                )
                db.add(member)
                await db.flush()
                members_by_name[name] = member
            else:
                member.role = role
                member.profile_img_url = image_url

            message = (
                await db.execute(
                    select(TeamMessage).where(TeamMessage.member_id == member.id)
                )
            ).scalar_one_or_none()

            if message is None:
                db.add(
                    TeamMessage(
                        member_id=member.id,
                        title=title,
                        content=content,
                        detail_img_url=image_url,
                    )
                )
            else:
                message.title = title
                message.content = content
                message.detail_img_url = image_url

            upserted += 1

        await db.commit()

    print(
        f"Done. upserted={upserted}, removed={removed}, "
        f"xlsx='{XLSX_PATH.name}', image_dir='{IMAGE_DIR.name}'"
    )


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Upsert team messages from 운영팀 한마디 response xlsx"
    )
    parser.add_argument(
        "--prune-missing",
        action="store_true",
        help="Delete team member/message rows not present in current xlsx",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    asyncio.run(seed_team_messages(prune_missing=args.prune_missing))
