"""
단풍바람 13기 메생결산 실데이터 삽입 스크립트.

- 명부(25-2 시트)   → User + Character(name=실명, detail_txt=닉네임, user_id) 생성
- 메생결산시트.xlsx  → Settlement 생성 (이름 기준으로 Character 매칭)

Usage:
    uv run python seed_real_data.py
"""

import asyncio
import datetime
from pathlib import Path

import openpyxl
from sqlalchemy import select

from database import async_session, init_db
from models.character import Character
from models.settlement import Settlement
from models.user import User

BASE_DIR = Path(__file__).parent
EXCEL_DIR = BASE_DIR / "13기 메생결산"
ROSTER_PATH = EXCEL_DIR / "25-2 단풍바람 명부.xlsx"
SETTLEMENT_PATH = EXCEL_DIR / "메생결산시트.xlsx"


def _parse_date(yymmdd: int) -> datetime.date:
    """YYMMDD 정수(예: 250916) → date(2025, 9, 16)."""
    s = f"{int(yymmdd):06d}"
    yy, mm, dd = int(s[0:2]), int(s[2:4]), int(s[4:6])
    year = 2000 + yy if yy < 50 else 1900 + yy
    return datetime.date(year, mm, dd)


def load_roster() -> list[dict]:
    """명부 25-2 시트에서 활성 회원 목록을 읽어 반환한다."""
    wb = openpyxl.load_workbook(ROSTER_PATH)
    ws = wb["25-2"]
    members: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, gender, _dept, student_id, nickname, level, server, _cgender, job = row[:9]
        if not name or str(gender).strip() == "탈퇴":
            continue
        if not nickname or not level or not server or not job:
            continue
        members.append(
            {
                "name": str(name).strip(),            # 실명 → User.name, Character.name
                "gender": "female" if str(gender).strip() == "여자" else "male",
                "student_id": str(int(student_id)) if student_id else None,
                "nickname": str(nickname).strip(),    # 닉네임 → Character.detail_txt
                "level": int(level),
                "server": str(server).strip(),
                "job": str(job).strip(),
            }
        )
    return members


def load_settlements() -> list[dict]:
    """메생결산시트 Sheet1에서 결산 항목을 읽어 반환한다."""
    wb = openpyxl.load_workbook(SETTLEMENT_PATH)
    ws = wb["Sheet1"]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, name, _nickname, date_int, img_name, caption = row[:6]
        if not name or not date_int:
            continue
        rows.append(
            {
                "name": str(name).strip(),
                "date": _parse_date(date_int),
                "img_name": str(img_name).strip() if img_name else None,
                "caption": str(caption).strip() if caption else "",
            }
        )
    return rows


def _resolve_img_ext(member_name: str, img_stem: str) -> str | None:
    """실제 파일 확장자를 찾아 img_url을 반환한다."""
    member_dir = EXCEL_DIR / member_name
    for ext in (".png", ".jpg", ".jpeg", ".PNG", ".JPG", ".JPEG"):
        if (member_dir / (img_stem + ext)).exists():
            return f"/static/settlements/{member_name}/{img_stem}{ext}"
    return None


async def seed() -> None:
    await init_db()

    roster = load_roster()
    settlement_rows = load_settlements()

    print(f"명부 회원 수: {len(roster)}명")
    print(f"결산 항목 수: {len(settlement_rows)}개\n")

    async with async_session() as db:

        # ── Phase 1: User 생성 ──────────────────────────────────────────
        name_to_user: dict[str, User] = {}

        for m in roster:
            r = await db.execute(select(User).where(User.name == m["name"]))
            user = r.scalar_one_or_none()
            if user is None:
                username = m["student_id"] or m["nickname"]
                # username 충돌 방어
                r2 = await db.execute(select(User).where(User.username == username))
                if r2.scalar_one_or_none() is not None:
                    username = m["nickname"]

                user = User(
                    username=username,
                    name=m["name"],
                    student_id=m["student_id"],
                    nickname=m["nickname"],
                    gender=m["gender"],
                )
                db.add(user)
                print(f"[USER+] {m['name']} (username={username})")
            else:
                print(f"[USER=] {m['name']} - 이미 존재")

            name_to_user[m["name"]] = user

        await db.flush()  # user.id 확보

        # ── Phase 2: Character 생성 (name=실명, detail_txt=닉네임, user_id) ──
        name_to_char: dict[str, Character] = {}

        for m in roster:
            r = await db.execute(select(Character).where(Character.name == m["name"]))
            char = r.scalar_one_or_none()
            user = name_to_user[m["name"]]

            if char is None:
                char = Character(
                    user_id=user.id,
                    name=m["name"],           # 실명
                    detail_txt=m["nickname"], # 닉네임
                    level=m["level"],
                    server=m["server"],
                    job=m["job"],
                )
                db.add(char)
                await db.flush()
                print(
                    f"[CHAR+] {m['name']} ({m['nickname']})"
                    f" Lv.{m['level']} {m['job']} / {m['server']}"
                )
            else:
                print(f"[CHAR=] {m['name']} - 이미 존재")

            name_to_char[m["name"]] = char

        await db.flush()

        # ── Phase 3: Settlement 생성 ────────────────────────────────────
        print()
        s_ok = 0
        s_skip = 0
        for s in settlement_rows:
            char = name_to_char.get(s["name"])
            if char is None:
                print(f"[WARN]  Settlement 건너뜀 (미매칭: {s['name']})")
                s_skip += 1
                continue

            # 이미지명에 쉼표가 있는 경우(복수 이미지) → 첫 번째만 처리
            img_stems = [x.strip() for x in s["img_name"].split(",")] if s["img_name"] else []

            for i, stem in enumerate(img_stems):
                img_url = _resolve_img_ext(s["name"], stem)
                db.add(
                    Settlement(
                        character_id=char.id,
                        title=s["caption"],
                        description=None,
                        img_url=img_url,
                        acquired_at=s["date"],
                    )
                )
                s_ok += 1

            if not img_stems:
                db.add(
                    Settlement(
                        character_id=char.id,
                        title=s["caption"],
                        description=None,
                        img_url=None,
                        acquired_at=s["date"],
                    )
                )
                s_ok += 1

        await db.commit()

    print(f"\n{'=' * 50}")
    print(f"✓ 완료!")
    print(f"  User: {len(roster)}명 처리")
    print(f"  Character: {len(name_to_char)}개 처리")
    print(f"  Settlement: {s_ok}개 삽입 (건너뜀: {s_skip}개)")
    print(f"{'=' * 50}")


if __name__ == "__main__":
    asyncio.run(seed())
